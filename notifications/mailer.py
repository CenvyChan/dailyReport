import logging
from io import BytesIO

from django.conf import settings
from django.core.mail import EmailMultiAlternatives, get_connection
from django.template.loader import render_to_string
from openpyxl import Workbook

from notifications.models import DeliveryLog
from notifications.reporting import build_daily_report, has_any_data


logger = logging.getLogger(__name__)


def build_workbook(report):
    """销售、采购各一个工作表，只放当天明细；汇总在正文里。"""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for key, title in (("sales", "销售明细"), ("purchase", "采购明细")):
        section = report[key]
        if section is None:
            continue
        sheet = workbook.create_sheet(title)
        sheet.append(section["headers"])
        for row in section["rows"]:
            sheet.append(row)
    if not workbook.sheetnames:
        workbook.create_sheet("无数据")
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def render_email(report):
    subject = f"[{report['company'].name}] {report['report_date']:%Y-%m-%d} 经营日报"
    context = {
        "report": report,
        "sales": report["sales"],
        "purchase": report["purchase"],
        "section_list": [
            {"title": title, "data": report[key]}
            for key, title in (("sales", "销售"), ("purchase", "采购"))
            if report[key] is not None
        ],
    }
    return subject, render_to_string("notifications/daily_report.txt", context), render_to_string(
        "notifications/daily_report.html", context
    )


def send_daily_report(mailing_list, report_date, *, connection=None, skip_when_empty=True):
    """发送一个收件组的当日邮件，成败都写 DeliveryLog；返回 (是否发送, 说明)。"""
    report = build_daily_report(
        company=mailing_list.company,
        report_date=report_date,
        include_sales=mailing_list.includes_sales,
        include_purchase=mailing_list.includes_purchase,
    )
    if skip_when_empty and not has_any_data(report):
        return False, "当天没有数据，已跳过"

    subject, text_body, html_body = render_email(report)
    message = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=mailing_list.recipient_list(),
        cc=mailing_list.cc_list(),
        connection=connection,
    )
    message.attach_alternative(html_body, "text/html")
    if mailing_list.attach_workbook:
        message.attach(
            f"{mailing_list.company.code}-daily-{report_date:%Y%m%d}.xlsx",
            build_workbook(report),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    try:
        message.send(fail_silently=False)
    except Exception as exc:
        DeliveryLog.objects.create(
            mailing_list=mailing_list,
            report_date=report_date,
            status=DeliveryLog.Status.FAILED,
            recipient_count=len(mailing_list.recipient_list()),
            subject=subject,
            message=f"{type(exc).__name__}: {exc}",
        )
        # 定时发信没人盯着终端，堆栈要落到 logs/mail.log 才查得到。
        logger.exception(
            "发送失败 company=%s list=%s date=%s",
            mailing_list.company.code,
            mailing_list.pk,
            report_date,
        )
        return False, f"发送失败：{exc}"

    DeliveryLog.objects.create(
        mailing_list=mailing_list,
        report_date=report_date,
        status=DeliveryLog.Status.SENT,
        recipient_count=len(mailing_list.recipient_list()),
        subject=subject,
    )
    logger.info(
        "发送成功 company=%s list=%s date=%s recipients=%s",
        mailing_list.company.code,
        mailing_list.pk,
        report_date,
        len(mailing_list.recipient_list()),
    )
    return True, f"已发送给 {len(mailing_list.recipient_list())} 个收件人"


def smtp_connection():
    return get_connection()
