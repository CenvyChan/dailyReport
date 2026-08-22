from datetime import date, time
from io import BytesIO

from django.contrib.auth.models import User
from django.core import mail
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from core.models import Customer
from core.testing import company_a, company_b
from notifications.mailer import send_daily_report
from notifications.models import DeliveryLog, MailingList
from sales.models import SalesShipment


def make_list(company, **overrides):
    payload = {
        "company": company,
        "name": "管理层日报",
        "recipients": "boss@example.com, cfo@example.com",
        "send_at": time(18, 0),
    }
    payload.update(overrides)
    return MailingList.objects.create(**payload)


@override_settings(
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    DEFAULT_FROM_EMAIL="report@example.com",
)
class SendDailyReportTests(TestCase):
    def setUp(self):
        self.company = company_a()
        self.owner = User.objects.create_user("sales-a")
        self.customer = Customer.objects.create(company=self.company, name="客户 A")
        SalesShipment.objects.create(
            company=self.company,
            customer=self.customer,
            owner=self.owner,
            sale_type="DOMESTIC",
            shipment_date=date(2026, 8, 21),
            quantity=3,
            currency="CNY",
            original_amount="120.00",
            exchange_rate="1.0000",
            amount_cny="120.00",
        )

    def test_email_goes_to_every_recipient_with_company_in_the_subject(self):
        mailing_list = make_list(self.company)

        ok, _ = send_daily_report(mailing_list, date(2026, 8, 21))

        self.assertTrue(ok)
        self.assertEqual(len(mail.outbox), 1)
        message = mail.outbox[0]
        self.assertEqual(message.to, ["boss@example.com", "cfo@example.com"])
        self.assertIn(self.company.name, message.subject)
        self.assertIn("2026-08-21", message.subject)

    def test_body_carries_both_plain_text_and_html_with_the_totals(self):
        mailing_list = make_list(self.company)

        send_daily_report(mailing_list, date(2026, 8, 21))

        message = mail.outbox[0]
        self.assertIn("120.00", message.body)
        html_body = next(content for content, mime in message.alternatives if mime == "text/html")
        self.assertIn("本月累计", html_body)
        self.assertIn("客户 A", html_body)

    def test_workbook_attachment_only_holds_the_current_company_rows(self):
        other = company_b()
        other_customer = Customer.objects.create(company=other, name="客户 B 家")
        SalesShipment.objects.create(
            company=other,
            customer=other_customer,
            owner=self.owner,
            sale_type="DOMESTIC",
            shipment_date=date(2026, 8, 21),
            quantity=9,
            currency="CNY",
            original_amount="900.00",
            exchange_rate="1.0000",
            amount_cny="900.00",
        )
        mailing_list = make_list(self.company)

        send_daily_report(mailing_list, date(2026, 8, 21))

        name, content, _ = mail.outbox[0].attachments[0]
        self.assertEqual(name, "A-daily-20260821.xlsx")
        sheet = load_workbook(BytesIO(content))["销售明细"]
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(row=2, column=2).value, "客户 A")

    def test_attachment_can_be_switched_off(self):
        mailing_list = make_list(self.company, attach_workbook=False)

        send_daily_report(mailing_list, date(2026, 8, 21))

        self.assertEqual(mail.outbox[0].attachments, [])

    def test_purchase_only_scope_leaves_sales_out_of_the_email(self):
        mailing_list = make_list(self.company, scope=MailingList.Scope.PURCHASE)

        ok, detail = send_daily_report(mailing_list, date(2026, 8, 21))

        self.assertFalse(ok)
        self.assertIn("没有数据", detail)
        self.assertEqual(mail.outbox, [])

    def test_a_day_without_data_is_skipped_and_leaves_no_log(self):
        mailing_list = make_list(self.company)

        ok, detail = send_daily_report(mailing_list, date(2026, 8, 20))

        self.assertFalse(ok)
        self.assertIn("没有数据", detail)
        self.assertEqual(mail.outbox, [])
        self.assertEqual(DeliveryLog.objects.count(), 0)

    def test_an_empty_day_is_still_sent_when_asked(self):
        mailing_list = make_list(self.company)

        ok, _ = send_daily_report(mailing_list, date(2026, 8, 20), skip_when_empty=False)

        self.assertTrue(ok)
        self.assertEqual(len(mail.outbox), 1)

    def test_success_is_recorded_in_the_delivery_log(self):
        mailing_list = make_list(self.company)

        send_daily_report(mailing_list, date(2026, 8, 21))

        log = DeliveryLog.objects.get()
        self.assertEqual(log.status, DeliveryLog.Status.SENT)
        self.assertEqual(log.report_date, date(2026, 8, 21))
        self.assertEqual(log.recipient_count, 2)


@override_settings(
    EMAIL_BACKEND="notifications.tests.test_mailer.BrokenEmailBackend",
    DEFAULT_FROM_EMAIL="report@example.com",
)
class SendFailureTests(TestCase):
    def setUp(self):
        self.company = company_a()
        owner = User.objects.create_user("sales-a")
        customer = Customer.objects.create(company=self.company, name="客户 A")
        SalesShipment.objects.create(
            company=self.company,
            customer=customer,
            owner=owner,
            sale_type="DOMESTIC",
            shipment_date=date(2026, 8, 21),
            quantity=1,
            currency="CNY",
            original_amount="10.00",
            exchange_rate="1.0000",
            amount_cny="10.00",
        )

    def test_smtp_failure_is_captured_in_the_log_instead_of_raising(self):
        mailing_list = make_list(self.company)

        ok, detail = send_daily_report(mailing_list, date(2026, 8, 21))

        self.assertFalse(ok)
        self.assertIn("发送失败", detail)
        log = DeliveryLog.objects.get()
        self.assertEqual(log.status, DeliveryLog.Status.FAILED)
        self.assertIn("SMTP 拒绝", log.message)


class BrokenEmailBackend:
    def __init__(self, *args, **kwargs):
        pass

    def send_messages(self, messages):
        raise RuntimeError("SMTP 拒绝了这次投递")

    def open(self):
        return True

    def close(self):
        pass
