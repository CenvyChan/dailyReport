from datetime import date
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from core.models import Customer, Supplier
from core.testing import company_a, company_b, login_with_company
from reports.exporters import sales_export_rows
from purchase.models import PurchaseReceipt
from sales.models import SalesShipment


class SalesExportTests(TestCase):
    def setUp(self):
        self.company = company_a()
        user = User.objects.create_user("sales-a")
        customer = Customer.objects.create(company=self.company, name="客户 A")
        SalesShipment.objects.create(
            company=self.company,
            customer=customer,
            owner=user,
            sale_type="DOMESTIC",
            shipment_date=date(2026, 8, 10),
            quantity=1,
            currency="CNY",
            original_amount="88.00",
            exchange_rate="1.0000",
            amount_cny="88.00",
        )

    def test_export_headers_include_original_and_converted_amounts(self):
        headers, rows = sales_export_rows(SalesShipment.objects.all())
        self.assertEqual(
            headers,
            ["出货日期", "客户", "负责人", "销售类型", "数量", "币种", "原币金额", "汇率快照", "折算人民币金额"],
        )
        # rows 是生成器（流式写出，避免整表物化），所以先收集再断言。
        self.assertEqual(len(list(rows)), 1)

    def test_sales_export_applies_dashboard_filters(self):
        admin = User.objects.create_superuser("admin", password="pw")
        customer = Customer.objects.get(company=self.company, name="客户 A")
        SalesShipment.objects.create(
            company=self.company,
            customer=customer,
            owner=admin,
            sale_type="DOMESTIC",
            shipment_date=date(2026, 8, 11),
            quantity=2,
            currency="CNY",
            original_amount="99.00",
            exchange_rate="1.0000",
            amount_cny="99.00",
        )
        login_with_company(self.client, admin, self.company)

        response = self.client.get(reverse("reports:sales_export"), {"start": "2026-08-11"})

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.max_row, 2)
        self.assertEqual(workbook.active.cell(row=2, column=1).value.date(), date(2026, 8, 11))

    def test_sales_export_only_contains_the_active_company(self):
        admin = User.objects.create_superuser("admin", password="pw")
        other = company_b()
        other_customer = Customer.objects.create(company=other, name="客户 B 家")
        SalesShipment.objects.create(
            company=other,
            customer=other_customer,
            owner=admin,
            sale_type="DOMESTIC",
            shipment_date=date(2026, 8, 10),
            quantity=7,
            currency="CNY",
            original_amount="777.00",
            exchange_rate="1.0000",
            amount_cny="777.00",
        )
        login_with_company(self.client, admin, self.company)

        response = self.client.get(reverse("reports:sales_export"))

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.max_row, 2)
        self.assertEqual(workbook.active.cell(row=2, column=2).value, "客户 A")

    def test_purchase_export_applies_dashboard_filters(self):
        admin = User.objects.create_superuser("admin", password="pw")
        supplier = Supplier.objects.create(company=self.company, name="供应商 A")
        for purchase_date, amount in ((date(2026, 8, 10), "88.00"), (date(2026, 8, 11), "99.00")):
            PurchaseReceipt.objects.create(
                company=self.company,
                supplier=supplier,
                buyer=admin,
                purchase_type="DOMESTIC",
                purchase_date=purchase_date,
                quantity=1,
                currency="CNY",
                original_amount=amount,
                exchange_rate="1.0000",
                amount_cny=amount,
            )
        login_with_company(self.client, admin, self.company)

        response = self.client.get(reverse("reports:purchase_export"), {"start": "2026-08-11"})

        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.max_row, 2)
        self.assertEqual(workbook.active.cell(row=2, column=1).value.date(), date(2026, 8, 11))


class ExportSizeLimitTests(TestCase):
    """导出无上限时，用户不设日期直接点一下就是全表。几十万行 × 9 列在
    Core2 Duo 上必然请求超时，用户只会反复点击。"""

    def setUp(self):
        self.admin = User.objects.create_superuser("admin", password="pw")
        self.company = company_a()
        customer = Customer.objects.create(company=self.company, name="客户 A")
        for day in (10, 11, 12):
            SalesShipment.objects.create(
                company=self.company,
                customer=customer,
                owner=self.admin,
                sale_type="DOMESTIC",
                shipment_date=date(2026, 8, day),
                quantity=1,
                currency="CNY",
                original_amount="10.00",
                exchange_rate="1.0000",
                amount_cny="10.00",
            )
        login_with_company(self.client, self.admin, self.company)

    def test_oversized_export_explains_how_to_narrow_the_range(self):
        with patch("reports.exporters.MAX_EXPORT_ROWS", 2):
            response = self.client.get(reverse("reports:sales_export"))

        self.assertEqual(response.status_code, 400)
        self.assertContains(response, "超过", status_code=400)
        self.assertContains(response, "缩小日期范围", status_code=400)
        # 必须是能看懂的页面，不是裸文本或 JSON。
        self.assertContains(response, "导出范围过大", status_code=400)

    def test_narrowing_the_range_lets_the_export_through(self):
        with patch("reports.exporters.MAX_EXPORT_ROWS", 2):
            response = self.client.get(reverse("reports:sales_export"), {"start": "2026-08-12"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(load_workbook(BytesIO(response.content)).active.max_row, 2)

    def test_row_count_within_the_limit_still_exports(self):
        response = self.client.get(reverse("reports:sales_export"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(load_workbook(BytesIO(response.content)).active.max_row, 4)

    def test_the_error_page_does_not_echo_an_external_referer(self):
        """返回链接来自 Referer，不能把外站地址渲染成本站按钮。"""
        with patch("reports.exporters.MAX_EXPORT_ROWS", 2):
            response = self.client.get(
                reverse("reports:sales_export"), HTTP_REFERER="https://evil.example.com/x"
            )

        self.assertNotContains(response, "evil.example.com", status_code=400)

    def test_export_message_reports_the_actual_row_count(self):
        with patch("reports.exporters.MAX_EXPORT_ROWS", 2):
            response = self.client.get(reverse("reports:sales_export"))

        self.assertContains(response, "3 行", status_code=400)
