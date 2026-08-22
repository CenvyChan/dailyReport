"""导入模板下载。列名必须与各 importer 实际读取的表头完全一致，
改这里的同时要改对应 importer，`core/tests/test_import_templates.py` 会校验两边一致。"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="DDEBF7")
HEADER_FONT = Font(bold=True)

TEMPLATES = {
    "customer": {
        "filename": "客户导入模板.xlsx",
        "sheets": [
            {
                "title": "客户",
                "headers": ["客户名称"],
                "samples": [["示例客户一"], ["示例客户二"]],
                "notes": [
                    "客户名称必填，请使用公司统一全称。",
                    "同名客户不会重复创建；已停用的同名客户会重新启用。",
                    "客户只会导入到你当前登录选择的公司。",
                ],
            }
        ],
    },
    "supplier": {
        "filename": "供应商导入模板.xlsx",
        "sheets": [
            {
                "title": "供应商",
                "headers": ["供应商名称"],
                "samples": [["示例供应商一"], ["示例供应商二"]],
                "notes": [
                    "供应商名称必填，请使用公司统一全称。",
                    "同名供应商不会重复创建；已停用的会重新启用。",
                    "供应商只会导入到你当前登录选择的公司。",
                ],
            }
        ],
    },
    "user": {
        "filename": "用户导入模板.xlsx",
        "sheets": [
            {
                "title": "用户",
                "headers": ["用户名", "姓名", "角色", "初始密码"],
                "samples": [
                    ["sales01", "张三", "销售", "Init@123456"],
                    ["both01", "李四", "销售、采购", "Init@123456"],
                ],
                "notes": [
                    "用户名、角色、初始密码必填；姓名可留空。",
                    "角色只能填：管理员、销售、采购、报表查看者。",
                    "一个人可以兼多个角色，用「、」或逗号分隔，例如「销售、采购」。",
                    "已存在的用户名不会被覆盖，会在预览中报错。",
                    "导入的账号会被授予当前登录公司的访问权限；要进入另一家公司，请在后台追加公司授权。",
                    "用户首次登录后必须自行修改密码。",
                ],
            }
        ],
    },
    "sales": {
        "filename": "销售历史数据导入模板.xlsx",
        "sheets": [
            {
                "title": "数据表",
                "headers": ["客户名称", "业务跟单", "销售类型", "出货日期", "数量", "金额"],
                "samples": [
                    ["示例客户一", "张三", "内销", "2026-08-10", 100, 12345.67],
                    ["示例客户二", "张三", "外销", "2026-08-11", 50, 8000],
                ],
                "notes": [
                    "工作表名称必须保持为「数据表」，不要改名。",
                    "销售类型只能填：内销、外销。内销按人民币，外销按美元。",
                    "出货日期用 2026-08-10 这种格式；数量填正整数，金额填原币金额。",
                    "业务跟单填销售业务员的用户名；系统会自动建立客户归属关系。",
                    "外销行所在月份必须在「汇率」表里有汇率，否则预览会报错。",
                ],
            },
            {
                "title": "汇率",
                "headers": ["日期", "汇率"],
                "samples": [["2026年8月", 7.12], ["2026年9月", 7.08]],
                "notes": [
                    "工作表名称必须保持为「汇率」，不要改名。",
                    "日期写成「2026年8月」这种中文格式；汇率填 1 美元兑多少人民币。",
                    "汇率会导入到当前登录公司，两家公司可以各维护一套。",
                ],
            },
        ],
    },
    "purchase": {
        "filename": "采购历史数据导入模板.xlsx",
        "sheets": [
            {
                "title": "数据表",
                "headers": ["供应商", "采购员", "采购类型", "采购日期", "数量", "金额"],
                "samples": [
                    ["示例供应商一", "李四", "国内采购", "2026-08-10", 200, 23456.78],
                    ["示例供应商二", "李四", "国外采购", "2026-08-11", 80, 9000],
                ],
                "notes": [
                    "采购类型只能填：国内采购、国外采购。国内按人民币，国外按美元。",
                    "采购日期用 2026-08-10 这种格式；数量填正整数，金额填原币金额。",
                    "采购员填采购人员的用户名；系统会自动建立供应商归属关系。",
                    "国外采购需要先在「汇率」维护好对应月份的汇率，否则预览会报错。",
                    "采购导入没有汇率工作表，请先在系统的汇率页面维护。",
                ],
            }
        ],
    },
}


NOTES_SHEET_TITLE = "填写说明"


def _write_sheet(sheet, spec):
    sheet.append(spec["headers"])
    for column_index, _ in enumerate(spec["headers"], start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(column_index)].width = 18
    for row in spec["samples"]:
        sheet.append(row)


def _write_notes(sheet, spec):
    """说明必须单独放一个工作表。放在数据表里会被 pandas 当成数据行，导入时整片报错。"""
    sheet.column_dimensions["A"].width = 90
    row = 1
    for sheet_spec in spec["sheets"]:
        sheet.cell(row=row, column=1, value=f"【{sheet_spec['title']}】").font = HEADER_FONT
        row += 1
        for index, note in enumerate(sheet_spec["notes"], start=1):
            sheet.cell(row=row, column=1, value=f"{index}. {note}")
            row += 1
        row += 1
    sheet.cell(row=row, column=1, value="示例行是给你照着改的，正式导入前请替换成真实数据或删除。").font = HEADER_FONT


def build_template(kind):
    spec = TEMPLATES[kind]
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_spec in spec["sheets"]:
        _write_sheet(workbook.create_sheet(sheet_spec["title"]), sheet_spec)
    _write_notes(workbook.create_sheet(NOTES_SHEET_TITLE), spec)
    buffer = BytesIO()
    workbook.save(buffer)
    return spec["filename"], buffer.getvalue()
