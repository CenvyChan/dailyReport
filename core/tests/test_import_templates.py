from io import BytesIO

from django.contrib.auth.models import User
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from core.importers import parse_roles, preview_customer_import, preview_supplier_import, preview_user_import
from core.templates_export import TEMPLATES, build_template
from core.testing import company_a, login_with_company
from purchase.importers import PURCHASE_COLUMNS, PURCHASE_TYPES, preview_purchase_import
from sales.importers import SALES_COLUMNS, SALE_TYPES


def sheet_of(kind, title):
    _, content = build_template(kind)
    return load_workbook(BytesIO(content))[title]


def headers_of(kind, title):
    sheet = sheet_of(kind, title)
    return [cell.value for cell in sheet[1] if cell.value is not None]


class TemplateHeaderContractTests(SimpleTestCase):
    """模板列名必须和 importer 实际读取的表头一致，否则用户照模板填也会导入失败。"""

    def test_sales_template_matches_the_columns_the_importer_requires(self):
        self.assertEqual(headers_of("sales", "数据表"), [aliases[0] for aliases in SALES_COLUMNS])

    def test_sales_template_has_the_exchange_rate_sheet_the_importer_reads(self):
        self.assertEqual(headers_of("sales", "汇率"), ["日期", "汇率"])

    def test_purchase_template_matches_the_primary_alias_of_each_column(self):
        self.assertEqual(headers_of("purchase", "数据表"), [aliases[0] for aliases in PURCHASE_COLUMNS])

    def test_user_template_matches_the_required_user_columns(self):
        self.assertEqual(headers_of("user", "用户"), ["用户名", "姓名", "角色", "初始密码"])

    def test_customer_and_supplier_templates_use_the_recognised_name_column(self):
        self.assertEqual(headers_of("customer", "客户"), ["客户名称"])
        self.assertEqual(headers_of("supplier", "供应商"), ["供应商名称"])

    def test_sample_business_types_are_values_the_importers_accept(self):
        sales_types = {row[2] for row in TEMPLATES["sales"]["sheets"][0]["samples"]}
        self.assertTrue(sales_types.issubset(SALE_TYPES))

        purchase_types = {row[2] for row in TEMPLATES["purchase"]["sheets"][0]["samples"]}
        self.assertTrue(purchase_types.issubset(PURCHASE_TYPES))

    def test_sample_roles_are_values_the_user_importer_accepts(self):
        for row in TEMPLATES["user"]["sheets"][0]["samples"]:
            with self.subTest(cell=row[2]):
                roles, invalid = parse_roles(row[2])
                self.assertIsNone(invalid)
                self.assertTrue(roles)

    def test_user_template_demonstrates_the_multi_role_syntax(self):
        cells = [row[2] for row in TEMPLATES["user"]["sheets"][0]["samples"]]

        self.assertTrue(any(len(parse_roles(cell)[0]) > 1 for cell in cells))

    def test_every_template_carries_samples_and_notes(self):
        for kind, spec in TEMPLATES.items():
            for sheet in spec["sheets"]:
                with self.subTest(kind=kind, sheet=sheet["title"]):
                    self.assertTrue(sheet["samples"])
                    self.assertTrue(sheet["notes"])
                    for row in sheet["samples"]:
                        self.assertEqual(len(row), len(sheet["headers"]))


class TemplateRoundTripTests(TestCase):
    """把模板原样喂回 importer，示例行应该被认成有效数据（说明列名和格式都对得上）。"""

    def test_customer_template_samples_validate_cleanly(self):
        _, content = build_template("customer")
        preview = preview_customer_import(BytesIO(content))

        self.assertEqual(preview.valid_row_count, 2)
        self.assertEqual(preview.error_rows, [])

    def test_supplier_template_samples_validate_cleanly(self):
        _, content = build_template("supplier")
        preview = preview_supplier_import(BytesIO(content))

        self.assertEqual(preview.valid_row_count, 2)
        self.assertEqual(preview.error_rows, [])

    def test_user_template_samples_validate_cleanly(self):
        _, content = build_template("user")
        preview = preview_user_import(BytesIO(content))

        self.assertEqual(preview.valid_row_count, 2)
        self.assertEqual(preview.error_rows, [])

    def test_purchase_template_samples_validate_apart_from_the_missing_rate(self):
        _, content = build_template("purchase")
        preview = preview_purchase_import(BytesIO(content), company=company_a())

        self.assertEqual(preview.valid_row_count, 2)
        self.assertEqual(preview.error_rows, [])
        # 国外采购示例行需要汇率，模板说明里已提示先维护汇率。
        self.assertEqual([error["field"] for error in preview.rate_errors], ["汇率"])


class TemplateDownloadViewTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser("admin", password="pw")
        login_with_company(self.client, self.admin, company_a())

    def test_administrator_can_download_every_template(self):
        for kind in TEMPLATES:
            with self.subTest(kind=kind):
                response = self.client.get(reverse("core:import_template", args=[kind]))

                self.assertEqual(response.status_code, 200)
                self.assertIn("spreadsheetml", response["Content-Type"])
                self.assertIn("attachment", response["Content-Disposition"])
                self.assertTrue(load_workbook(BytesIO(response.content)).sheetnames)

    def test_unknown_template_is_a_404(self):
        self.assertEqual(self.client.get(reverse("core:import_template", args=["nope"])).status_code, 404)

    def test_non_administrator_cannot_download(self):
        user = User.objects.create_user("sales-a")
        login_with_company(self.client, user, company_a())

        self.assertEqual(self.client.get(reverse("core:import_template", args=["sales"])).status_code, 403)

    def test_import_pages_link_to_their_template(self):
        for url_name, kind in (
            ("core:customer_import_page", "customer"),
            ("core:supplier_import_page", "supplier"),
            ("core:user_import_page", "user"),
            ("sales:import_page", "sales"),
            ("purchase:import_page", "purchase"),
        ):
            with self.subTest(url_name=url_name):
                response = self.client.get(reverse(url_name))

                self.assertContains(response, reverse("core:import_template", args=[kind]))
                self.assertContains(response, "下载导入模板")
