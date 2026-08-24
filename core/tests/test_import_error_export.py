"""导入错误清单导出。

只在网页上列出错误时，用户得对着行号回 Excel 里逐条翻找——几百行历史数据里
错几行就很痛苦。这些测试锁住「原始内容和错误原因并排出现」这个核心价值。
"""

from io import BytesIO

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from core.import_errors import build_error_workbook
from core.testing import company_a, login_with_company
from core.tests.test_excel_formats import XLSX_MIME, xlsx_bytes


class BuildErrorWorkbookTests(SimpleTestCase):
    def test_original_row_sits_next_to_the_reason(self):
        """核心价值：用户能看到自己当时填的是什么，不用回原文件对行号。"""
        rows = [
            {"客户名称": "客户甲", "数量": 5},
            {"客户名称": "客户乙", "数量": -1},
        ]
        errors = [{"row_number": 3, "field": "数量", "message": "格式或数值无效"}]

        sheet = load_workbook(BytesIO(build_error_workbook([("错误清单", rows, errors)]))).active

        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["Excel 行号", "出错的列", "错误原因", "客户名称", "数量"],
        )
        # row_number=3 对应 rows[1]，也就是出错的那一行。
        self.assertEqual([cell.value for cell in sheet[2]], [3, "数量", "格式或数值无效", "客户乙", -1])

    def test_only_failing_rows_are_listed(self):
        rows = [{"客户名称": "甲"}, {"客户名称": "乙"}, {"客户名称": "丙"}]
        errors = [{"row_number": 4, "field": "客户名称", "message": "不能为空"}]

        sheet = load_workbook(BytesIO(build_error_workbook([("错误清单", rows, errors)]))).active

        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(row=2, column=4).value, "丙")

    def test_rate_errors_go_to_their_own_sheet(self):
        """销售导入的汇率错误来自「汇率」工作表，行号对的是那张表。
        混进数据表的清单里会对错行。"""
        data_rows = [{"客户名称": "甲"}]
        rate_rows = [{"月份": "2026-08", "汇率": "abc"}]

        workbook = load_workbook(
            BytesIO(
                build_error_workbook(
                    [
                        ("数据表", data_rows, [{"row_number": 2, "field": "客户名称", "message": "不能为空"}]),
                        ("汇率", rate_rows, [{"row_number": 2, "field": "汇率", "message": "格式无效"}]),
                    ]
                )
            )
        )

        self.assertEqual(workbook.sheetnames, ["数据表", "汇率"])
        self.assertEqual(workbook["数据表"].cell(row=2, column=4).value, "甲")
        self.assertEqual(workbook["汇率"].cell(row=2, column=5).value, "abc")

    def test_sections_without_errors_are_skipped(self):
        workbook = load_workbook(
            BytesIO(
                build_error_workbook(
                    [
                        ("数据表", [{"客户名称": "甲"}], [{"row_number": 2, "message": "不能为空"}]),
                        ("汇率", [{"汇率": 7.1}], []),
                    ]
                )
            )
        )

        self.assertEqual(workbook.sheetnames, ["数据表"])

    def test_a_row_number_beyond_the_data_does_not_break_the_export(self):
        """汇率类错误可能没有对应的数据行，越界时留空而不是整个导出失败。"""
        sheet = load_workbook(
            BytesIO(build_error_workbook([("错误清单", [{"客户名称": "甲"}], [{"row_number": 99, "message": "错"}])]))
        ).active

        self.assertEqual(sheet.cell(row=2, column=1).value, 99)
        self.assertIsNone(sheet.cell(row=2, column=4).value)

    def test_errors_without_a_row_number_still_appear(self):
        sheet = load_workbook(
            BytesIO(build_error_workbook([("错误清单", [], [{"message": "整张表都不对"}])]))
        ).active

        self.assertEqual(sheet.cell(row=2, column=3).value, "整张表都不对")

    def test_unnamed_columns_are_dropped(self):
        """pandas 把空列名读成 Unnamed: N，那种列对用户没意义。"""
        rows = [{"客户名称": "甲", "Unnamed: 3": None}]

        sheet = load_workbook(
            BytesIO(build_error_workbook([("错误清单", rows, [{"row_number": 2, "message": "错"}])]))
        ).active

        self.assertEqual([cell.value for cell in sheet[1]], ["Excel 行号", "出错的列", "错误原因", "客户名称"])

    def test_nan_becomes_blank_rather_than_breaking_the_save(self):
        """openpyxl 写不了 pandas 的 NaN。"""
        rows = [{"客户名称": "甲", "数量": float("nan")}]

        sheet = load_workbook(
            BytesIO(build_error_workbook([("错误清单", rows, [{"row_number": 2, "message": "错"}])]))
        ).active

        self.assertIsNone(sheet.cell(row=2, column=5).value)

    def test_the_header_row_is_frozen(self):
        """清单可能几十行，不冻结就得来回滚。"""
        sheet = load_workbook(
            BytesIO(build_error_workbook([("错误清单", [{"客户名称": "甲"}], [{"row_number": 2, "message": "错"}])]))
        ).active

        self.assertEqual(sheet.freeze_panes, "A2")

    def test_an_empty_error_list_still_produces_a_valid_file(self):
        sheet = load_workbook(BytesIO(build_error_workbook([("错误清单", [], [])]))).active

        self.assertEqual([cell.value for cell in sheet[1]], ["Excel 行号", "出错的列", "错误原因"])


class SalesErrorExportViewTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser("admin", password="pw")
        login_with_company(self.client, self.admin, company_a())

    def _upload(self, content, name="销售.xlsx"):
        return self.client.post(
            reverse("sales:import_errors_export"),
            {"file": SimpleUploadedFile(name, content, content_type=XLSX_MIME)},
        )

    def test_export_returns_an_xlsx_with_the_offending_row(self):
        content = xlsx_bytes(
            {
                "数据表": [
                    ["客户名称", "业务跟单", "销售类型", "出货日期", "数量", "金额"],
                    ["客户甲", "张三", "内销", "2026-08-10", 1, 100],
                    ["客户乙", "张三", "内销", "2026-08-11", -5, 100],
                ],
                "汇率": [["月份", "美元"], ["2026-08", 7.1]],
            }
        )

        response = self._upload(content)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_MIME)
        sheet = load_workbook(BytesIO(response.content))["数据表"]
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(row=2, column=1).value, 3)
        self.assertIn("客户乙", [cell.value for cell in sheet[2]])

    def test_the_download_gets_a_readable_chinese_filename(self):
        content = xlsx_bytes(
            {
                "数据表": [["客户名称", "业务跟单", "销售类型", "出货日期", "数量", "金额"], ["甲", "张三", "内销", "2026-08-10", -1, 1]],
                "汇率": [["月份", "美元"], ["2026-08", 7.1]],
            }
        )

        response = self._upload(content)

        # 中文名必须用 filename*=UTF-8'' 传，否则部分浏览器上会乱码。
        self.assertIn("filename*=UTF-8''", response["Content-Disposition"])

    def test_a_missing_file_is_rejected_with_a_readable_message(self):
        response = self.client.post(reverse("sales:import_errors_export"), {})

        self.assertEqual(response.status_code, 400)
        self.assertIn("请选择", response.json()["error"])

    def test_a_non_administrator_cannot_export(self):
        self.client.logout()
        staff = User.objects.create_user("someone", password="pw")
        login_with_company(self.client, staff, company_a())

        response = self.client.post(reverse("sales:import_errors_export"), {})

        self.assertEqual(response.status_code, 403)


class PurchaseErrorExportViewTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser("admin", password="pw")
        login_with_company(self.client, self.admin, company_a())

    def test_export_lists_the_offending_purchase_row(self):
        content = xlsx_bytes(
            {
                "Sheet1": [
                    ["供应商名称", "采购跟单", "采购类型", "入库日期", "数量", "金额", "币种"],
                    ["供应商甲", "李四", "内购", "2026-08-10", -3, 100, "CNY"],
                ]
            }
        )

        response = self.client.post(
            reverse("purchase:import_errors_export"),
            {"file": SimpleUploadedFile("采购.xlsx", content, content_type=XLSX_MIME)},
        )

        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.content)).active
        self.assertIn("供应商甲", [cell.value for cell in sheet[2]])
