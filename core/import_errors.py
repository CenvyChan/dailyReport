"""把导入校验的错误清单导出成 Excel。

只在网页上列出错误行时，用户得对着行号回 Excel 里逐条翻找——几百行历史数据
里错几行就很痛苦。导出成表格后，原始内容和错误原因并排放，改完直接重传。

行号与 validator 里的 enumerate(rows, start=2) 一致：第 1 行是表头，数据从
第 2 行开始，所以 rows[row_number - 2] 就是对应的原始行。
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROW_NUMBER_HEADER = "Excel 行号"
FIELD_HEADER = "出错的列"
MESSAGE_HEADER = "错误原因"
DATA_START_ROW = 2


def build_error_workbook(sections):
    """sections 是 [(工作表名, 原始行, 错误清单), ...]。

    分成多个工作表是必要的：销售导入的汇率错误来自「汇率」工作表，行号对的是
    那张表，混在数据表的清单里会对错行。

    每个工作表的表头是「行号 + 出错的列 + 错误原因 + 原始各列」，用户既能看到
    问题，也能看到自己当时填的内容。
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    for title, rows, errors in sections:
        if not errors:
            continue
        sheet = workbook.create_sheet(_safe_title(title))
        source_columns = _source_columns(rows)
        headers = [ROW_NUMBER_HEADER, FIELD_HEADER, MESSAGE_HEADER, *source_columns]
        sheet.append(headers)
        for error in errors:
            row_number = error.get("row_number")
            original = _original_row(rows, row_number)
            sheet.append(
                [
                    row_number if row_number is not None else "",
                    error.get("field") or "",
                    error.get("message") or "",
                    *[_cell(original.get(column)) for column in source_columns],
                ]
            )
        _style(sheet, len(headers))

    if not workbook.sheetnames:
        # 没有错误也要给一个合法的工作簿，否则 openpyxl 存不出来。
        sheet = workbook.create_sheet("错误清单")
        sheet.append([ROW_NUMBER_HEADER, FIELD_HEADER, MESSAGE_HEADER])
        _style(sheet, 3)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _safe_title(title):
    """Excel 工作表名上限 31 字符，且不允许 : \\ / ? * [ ]。"""
    cleaned = "".join("_" if character in r":\/?*[]" else character for character in title)
    return cleaned[:31] or "错误清单"


def _source_columns(rows):
    """取原始表头。pandas 会把空列名读成 Unnamed: N，那种列对用户没意义，去掉。"""
    if not rows:
        return []
    return [
        str(key)
        for key in rows[0].keys()
        if not str(key).startswith("Unnamed:")
    ]


def _original_row(rows, row_number):
    """按行号取回原始行。行号越界或缺失时返回空 dict，不让导出整体失败——
    汇率类错误就没有对应的数据行。"""
    if row_number is None:
        return {}
    index = row_number - DATA_START_ROW
    if 0 <= index < len(rows):
        return rows[index]
    return {}


def _cell(value):
    """openpyxl 写不了 pandas 的 NaN 和 Timestamp，统一成字符串。
    NaN 自身不等于自身，用这个特性判空，不必引入 pandas。"""
    if value is None or (isinstance(value, float) and value != value):
        return ""
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def _style(sheet, column_count):
    """给表头加底色并冻结，列宽按内容估。清单可能几十行，不冻结就得来回滚。"""
    fill = PatternFill("solid", fgColor="FFF2CC")
    for column in range(1, column_count + 1):
        cell = sheet.cell(row=1, column=column)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"

    for column in range(1, column_count + 1):
        letter = get_column_letter(column)
        longest = max(
            (len(str(sheet.cell(row=row, column=column).value or "")) for row in range(1, sheet.max_row + 1)),
            default=0,
        )
        # 中文字符在 Excel 里约占两个字符宽，留些余量；上限防止某列过长把表撑开。
        sheet.column_dimensions[letter].width = min(max(longest + 4, 10), 40)
